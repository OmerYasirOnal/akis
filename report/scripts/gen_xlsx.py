from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

doc = Document("/tmp/report/build/working.docx")
wb = Workbook(); wb.remove(wb.active)

HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="BBBBBB"); BORDER = Border(thin,thin,thin,thin)

def add_sheet(name, title, rows, widths):
    ws = wb.create_sheet(name[:31])
    ws["A1"]=title; ws["A1"].font=Font(bold=True, size=12, color="1F3864")
    start=3
    for j,h in enumerate(rows[0],1):
        c=ws.cell(start,j,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=WRAP; c.border=BORDER
    for i,r in enumerate(rows[1:],start+1):
        for j,v in enumerate(r,1):
            c=ws.cell(i,j,v); c.alignment=WRAP; c.border=BORDER
    for j,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(j)].width=w
    return ws

# Map docx tables -> sheets (table order in doc: 5.1,5.2,8.1,8.2,8.3,9.1)
def tbl_rows(t):
    out=[]
    for row in t.rows:
        out.append([c.text.strip().replace("\n"," ") for c in row.cells])
    # de-dup horizontally-merged repeated cells per row
    cleaned=[]
    for r in out:
        nr=[]; prev=None
        for v in r:
            if v==prev: nr.append("")
            else: nr.append(v); prev=v
        cleaned.append(nr)
    return out  # keep raw; merges fine for data file

T = doc.tables
add_sheet("Table 5.1 FR","Table 5.1. Functional Requirements", tbl_rows(T[0]), [14,32,60])
add_sheet("Table 5.2 NFR","Table 5.2. Non-Functional Requirements", tbl_rows(T[1]), [14,24,60])
add_sheet("Table 8.1 Tests","Table 8.1. Automated test execution results", tbl_rows(T[2]), [34,14])
add_sheet("Table 8.2 Coverage","Table 8.2. Frontend coverage results", tbl_rows(T[3]), [22,14])
add_sheet("Table 8.3 Mock","Table 8.3. Mock pipeline baseline evaluation scenarios", tbl_rows(T[4]), [28,14,18])
add_sheet("Table 9.1 Objectives","Table 9.1. Objectives vs achieved results", tbl_rows(T[5]), [44,44])

# ---- Chart data sheets ----
add_sheet("Fig 4.1 SWE-bench","Figure 4.1 data — resolved real GitHub issues (%)",
    [["System / benchmark","Resolved (%)","Source"],
     ["Best LLM (Claude 2), SWE-bench",1.96,"[16] Jimenez et al., 2024"],
     ["SWE-agent + GPT-4, SWE-bench",12.47,"[4] Yang et al., 2024"],
     ["SWE-agent + GPT-4, de-leaked (SWE-Bench+)",3.97,"[17] Aleithan et al., 2024"]], [40,14,28])
add_sheet("Fig 4.2 SelfCorrect","Figure 4.2 data — accuracy before/after intrinsic self-correction (%) [18]",
    [["Model / dataset","Standard","Self-correct r1","Self-correct r2"],
     ["GPT-3.5 / GSM8K",75.9,75.1,74.7],
     ["GPT-3.5 / CommonSenseQA",75.8,38.1,41.8],
     ["GPT-4 / GSM8K",95.5,91.5,89.0]], [26,12,16,16])
add_sheet("Fig 8.3 Tests","Figure 8.3 data — automated test execution",
    [["Component","Passed","Skipped","Failed"],
     ["Backend",1285,5,0],["Frontend",415,0,0],["Total",1700,5,0]], [16,12,12,12])
add_sheet("Fig 8.4 Coverage","Figure 8.4 data — frontend coverage (%)",
    [["Metric","Coverage (%)"],["Statements",89.95],["Lines",89.95],["Branches",86.94],["Functions",80.51]], [18,16])
add_sheet("Fig 9.1 Objectives","Figure 9.1 data — objectives achievement",
    [["Status","Count"],["Implemented",10],["Not implemented",0]], [20,12])

wb.save("/tmp/report/build/AKIS_report_data.xlsx")
print("xlsx sheets:", [ws.title for ws in wb.worksheets])
